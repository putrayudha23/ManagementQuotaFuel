from django.shortcuts import render
from django.core.paginator import Paginator
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
import os
from django.conf import settings
from django.db.models import Q

from dataBph.models import NonTransportasiBPHMigas, ViewNonTransportasi
from userSetting.models import UserSiteMapping

# Create your views here.
def index(request):
    if request.user.is_authenticated:
        q1 = Q(status=2)
        q2 = Q(status=4)
        result = q1 | q2

        # get site user
        username = request.user.username
        user_item_mappings = UserSiteMapping.objects.filter(username=username)
        site_registrations = list(user_item_mappings.values_list('site_registration', flat=True))
        
        data = ViewNonTransportasi.objects.filter(data_release=None, site_registration__in=site_registrations).filter(result)

        paginator = Paginator(data, 100)  # Show 10 items per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context = {'page_obj': page_obj}

        if request.user.is_superuser:
            return render(request, 'dataBph.html', context)
        else:
            return render(request, 'dataBph_cabang.html', context)
    else:
        return render(request, 'noAccess.html')
    
def download_data(request):
    q1 = Q(status=2)
    q2 = Q(status=4)
    result = q1 | q2

    # get site user
    username = request.user.username
    user_item_mappings = UserSiteMapping.objects.filter(username=username)
    site_registrations = list(user_item_mappings.values_list('site_registration', flat=True))
    
    data = ViewNonTransportasi.objects.filter(data_release=None, site_registration__in=site_registrations).filter(result)
    
    # Load the existing workbook
    base_dir = settings.BASE_DIR
    excel_file_path = os.path.join(base_dir, 'nonVehicleMasterBph.xlsx')
    workbook = openpyxl.load_workbook(excel_file_path)
    
    # Access the active sheet (assuming it's the first sheet)
    sheet = workbook.active
    
    # Insert data into specific columns starting from the second row (row 2)
    for row_num, entry in enumerate(data, start=2):
        # Unhide the row if it was hidden
        sheet.row_dimensions[row_num].hidden = False
        
        sheet.cell(row=row_num, column=19).value = entry.badan_usaha
        sheet.cell(row=row_num, column=18).value = entry.no_surat_rekomendasi
        sheet.cell(row=row_num, column=4).value = entry.nik
        sheet.cell(row=row_num, column=3).value = entry.nama
        sheet.cell(row=row_num, column=20).value = entry.nama_usaha
        sheet.cell(row=row_num, column=21).value = entry.jenis_bbm_desc #/////
        sheet.cell(row=row_num, column=6).value = entry.sektor_konsumen_desc #/////
        sheet.cell(row=row_num, column=22).value = entry.jenis_usaha_desc #/////
        sheet.cell(row=row_num, column=7).value = entry.nama_kapal
        sheet.cell(row=row_num, column=15).value = entry.alokasi_volume
        sheet.cell(row=row_num, column=1).value = entry.site_registration
        sheet.cell(row=row_num, column=16).value = entry.tanggal_terbit
        sheet.cell(row=row_num, column=17).value = entry.tanggal_berakhir
        sheet.cell(row=row_num, column=23).value = entry.kode_provinsi
        sheet.cell(row=row_num, column=24).value = entry.kode_kabkota
        sheet.cell(row=row_num, column=25).value = entry.kode_kecamatan
        sheet.cell(row=row_num, column=26).value = entry.kode_keldesa
        # sheet.cell(row=row_num, column=).value = entry.created_date
        # sheet.cell(row=row_num, column=).value = entry.updated_date
        # sheet.cell(row=row_num, column=).value = entry.status
        sheet.cell(row=row_num, column=27).value = entry.penerbit
        # sheet.cell(row=row_num, column=).value = entry.url_file

    # Create a new file name to avoid overwriting the original
    new_file_name = 'new_nonVehicleMaster.xlsx'
    
    # Save the workbook with a new name
    workbook.save(new_file_name)
    
    # Prepare the response to download the edited file
    with open(new_file_name, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=' + new_file_name

    # Delete the file after serving the response
    if os.path.exists(new_file_name):
        os.remove(new_file_name)
    
    return response